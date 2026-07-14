# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
import calendar
import json
import os
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, send_file, Response

from .auth import login_required
from .db import get_db, get_open_cash_session_id
from .utils import parse_brl_to_cents, cents_to_brl, today_yyyy_mm_dd

bp = Blueprint("finance", __name__, url_prefix="/finance")

PAYMENT_METHODS = [
    ("cash", "Dinheiro"),
    ("pix", "Pix"),
    ("card_credit", "Cartão crédito"),
    ("card_debit", "Cartão débito"),
    ("card", "Cartão"),  # legado
    ("transfer", "Transferência"),
    ("boleto", "Boleto bancário"),
    ("other", "Outro"),
]
PAYMENT_METHOD_KEYS = {k for k, _ in PAYMENT_METHODS}


def app_setting(key: str, default: str = "") -> str:
    try:
        row = get_db().execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        return (row["value"] if row and row["value"] is not None else default)
    except Exception:
        return default



def finance_required(view):
    """Proteção extra do módulo Financeiro por senha (além do login)."""

    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("finance_unlocked"):
            nxt = request.full_path if request.query_string else request.path
            return redirect(url_for("finance.unlock", next=nxt))
        return view(*args, **kwargs)

    return wrapped


def normalize_payment_method(method: str | None) -> str:
    method = (method or "pix").strip()
    if method not in PAYMENT_METHOD_KEYS:
        return "other"
    return method


def parse_percent(value: str | None) -> float:
    s = (value or "0").strip().replace("%", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        val = float(Decimal(s))
    except (InvalidOperation, ValueError):
        return 0.0
    return max(0.0, min(100.0, val))


def fmt_percent(value: float | int | None) -> str:
    try:
        d = Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if d == d.to_integral():
            return str(int(d))
        return str(d).replace(".", ",")
    except Exception:
        return "0"


def calc_final_amount(gross_cents: int, discount_percent: float, discount_cents: int) -> tuple[int, int]:
    gross_cents = max(0, int(gross_cents or 0))
    discount_cents = max(0, int(discount_cents or 0))
    percent_cents = int((Decimal(gross_cents) * Decimal(str(discount_percent)) / Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    total_discount = min(gross_cents, percent_cents + discount_cents)
    return gross_cents - total_discount, total_discount


def add_months(date_str: str, months: int) -> str:
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        d = date.today()
    month_index = d.month - 1 + months
    year = d.year + month_index // 12
    month = month_index % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat()


def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def sync_transaction_payments(db, tid: int) -> None:
    tx = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    if not tx:
        return
    final_amount = int(tx["final_amount_cents"] if "final_amount_cents" in tx.keys() and tx["final_amount_cents"] is not None else tx["amount_cents"] or 0)
    paid = db.execute("SELECT COALESCE(SUM(amount_cents), 0) AS s FROM transaction_payments WHERE transaction_id=?", (tid,)).fetchone()["s"]
    paid = int(paid or 0)
    balance = max(final_amount - paid, 0)
    status = "paid" if final_amount > 0 and paid >= final_amount else "pending"

    last_payment = db.execute(
        "SELECT payment_method, date, cash_session_id FROM transaction_payments WHERE transaction_id=? ORDER BY date DESC, id DESC LIMIT 1",
        (tid,),
    ).fetchone()
    payment_method = tx["payment_method"]
    tx_date = tx["date"]
    cash_session_id = tx["cash_session_id"]
    if last_payment:
        payment_method = last_payment["payment_method"] or payment_method
        if status == "paid":
            tx_date = last_payment["date"] or tx_date
        cash_session_id = last_payment["cash_session_id"] if last_payment["cash_session_id"] is not None else cash_session_id

    db.execute(
        "UPDATE transactions SET paid_amount_cents=?, balance_cents=?, status=?, amount_cents=?, payment_method=?, date=?, cash_session_id=? WHERE id=?",
        (paid, balance, status, final_amount, payment_method, tx_date, cash_session_id, tid),
    )


def add_payment(db, tid: int, amount_cents: int, payment_method: str, pay_date: str, notes: str = "") -> bool:
    tx = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    if not tx:
        return False
    amount_cents = int(amount_cents or 0)
    if amount_cents <= 0:
        return False
    final_amount = int(tx["final_amount_cents"] or tx["amount_cents"] or 0)
    already_paid = int(tx["paid_amount_cents"] or 0)
    remaining = max(final_amount - already_paid, 0)
    if remaining > 0:
        amount_cents = min(amount_cents, remaining)
    payment_method = normalize_payment_method(payment_method)
    pay_date = pay_date or today_yyyy_mm_dd()

    cash_session_id = None
    open_cash_id = get_open_cash_session_id()
    if payment_method == "cash" and open_cash_id:
        cash_session_id = open_cash_id

    db.execute(
        "INSERT INTO transaction_payments(transaction_id, date, amount_cents, payment_method, notes, cash_session_id) VALUES(?,?,?,?,?,?)",
        (tid, pay_date, amount_cents, payment_method, notes.strip() or None, cash_session_id),
    )
    sync_transaction_payments(db, tid)
    return True


def get_common_form_data(db):
    return {
        "patients": db.execute("SELECT id, name FROM patients ORDER BY name ASC").fetchall(),
        "categories": db.execute("SELECT id, name, kind FROM categories WHERE active=1 ORDER BY name ASC").fetchall(),
        "providers": db.execute("SELECT id, name, default_repasse_percent FROM providers WHERE active=1 ORDER BY name ASC").fetchall(),
        "plan_items": db.execute(
            "SELECT pi.id, pi.patient_id, pi.procedure, pi.amount_cents, pi.done, p.name AS patient_name "
            "FROM plan_items pi JOIN patients p ON p.id=pi.patient_id "
            "ORDER BY p.name COLLATE NOCASE, pi.created_at DESC, pi.id DESC"
        ).fetchall(),
        "pm": PAYMENT_METHODS,
        "cents_to_brl": cents_to_brl,
    }


def get_filters():
    return {
        "kind": request.args.get("kind", "").strip(),
        "status": request.args.get("status", "").strip(),
        "payment_method": request.args.get("payment_method", "").strip(),
        "q": request.args.get("q", "").strip(),
        "date_from": request.args.get("from", "").strip(),
        "date_to": request.args.get("to", "").strip(),
        "patient_id": request.args.get("patient_id", "").strip(),
        "category_id": request.args.get("category_id", "").strip(),
        "provider_id": request.args.get("provider_id", "").strip(),
    }


def build_transactions_query(filters: dict, prefix: str = "t"):
    where = []
    params = []
    kind = filters.get("kind", "")
    status = filters.get("status", "")
    payment_method = filters.get("payment_method", "")
    q = filters.get("q", "")
    date_from = filters.get("date_from", "")
    date_to = filters.get("date_to", "")
    patient_id = filters.get("patient_id", "")
    category_id = filters.get("category_id", "")
    provider_id = filters.get("provider_id", "")

    if kind in ("income", "expense"):
        where.append(f"{prefix}.kind=?")
        params.append(kind)
    if status in ("paid", "pending"):
        where.append(f"{prefix}.status=?")
        params.append(status)
    elif status == "overdue":
        where.append(f"{prefix}.status='pending' AND COALESCE({prefix}.due_date, {prefix}.date) < ?")
        params.append(today_yyyy_mm_dd())
    if payment_method in PAYMENT_METHOD_KEYS:
        if payment_method == "card_credit":
            where.append(f"({prefix}.payment_method=? OR {prefix}.payment_method='card')")
            params.append("card_credit")
        else:
            where.append(f"{prefix}.payment_method=?")
            params.append(payment_method)
    if q:
        where.append(f"({prefix}.description LIKE ?)")
        params.append(f"%{q}%")
    if date_from:
        where.append(f"COALESCE({prefix}.due_date, {prefix}.date)>=?")
        params.append(date_from)
    if date_to:
        where.append(f"COALESCE({prefix}.due_date, {prefix}.date)<=?")
        params.append(date_to)
    if str(patient_id).isdigit():
        where.append(f"{prefix}.patient_id=?")
        params.append(int(patient_id))
    if str(category_id).isdigit():
        where.append(f"{prefix}.category_id=?")
        params.append(int(category_id))
    if str(provider_id).isdigit():
        where.append(f"{prefix}.provider_id=?")
        params.append(int(provider_id))

    return (("WHERE " + " AND ".join(where)) if where else ""), params


def insert_transaction(db, *, kind, status, date_eff, due_date, gross_amount, discount_percent, discount_fixed, description, pid, cid, prid, repasse_percent_int, payment_method, installments_total=1, installment_number=1, parent_transaction_id=None, plan_item_id=None):
    final_amount, total_discount = calc_final_amount(gross_amount, discount_percent, discount_fixed)
    cash_session_id = None
    db.execute(
        """
        INSERT INTO transactions(
            kind,status,date,due_date,amount_cents,payment_method,description,patient_id,category_id,provider_id,
            repasse_percent,cash_session_id,gross_amount_cents,discount_percent,discount_cents,final_amount_cents,
            paid_amount_cents,balance_cents,installments_total,installment_number,parent_transaction_id,plan_item_id
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            kind, "pending", date_eff, due_date, final_amount, payment_method, description, pid, cid, prid,
            repasse_percent_int, cash_session_id, gross_amount, discount_percent, total_discount, final_amount,
            0, final_amount, installments_total, installment_number, parent_transaction_id, plan_item_id,
        ),
    )
    tid = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    if status == "paid":
        add_payment(db, tid, final_amount, payment_method, date_eff, "Pagamento inicial")
    else:
        sync_transaction_payments(db, tid)
    return tid


@bp.route("/unlock", methods=["GET", "POST"])
@login_required
def unlock():
    next_url = request.args.get("next") or request.form.get("next") or url_for("finance.transactions")
    if request.method == "POST":
        senha = request.form.get("password", "")
        expected = (current_app.config.get("FINANCE_PASSWORD") or app_setting("finance_password", "") or "")
        if senha == expected:
            session["finance_unlocked"] = True
            flash("Financeiro desbloqueado ✅", "success")
            return redirect(next_url)
        flash("Senha do financeiro incorreta.", "danger")
    return render_template("finance_unlock.html", next=next_url)


@bp.route("/lock")
@login_required
def lock():
    session.pop("finance_unlocked", None)
    flash("Financeiro bloqueado 🔒", "info")
    return redirect(url_for("dashboard.index"))


@bp.route("/transactions")
@login_required
@finance_required
def transactions():
    db = get_db()
    filters = get_filters()
    where_sql, params = build_transactions_query(filters)
    rows = db.execute(
        "SELECT t.*, p.name AS patient_name, c.name AS category_name, pr.name AS provider_name, "
        "pi.procedure AS plan_item_name, pi.amount_cents AS plan_item_amount_cents "
        "FROM transactions t "
        "LEFT JOIN patients p ON p.id=t.patient_id "
        "LEFT JOIN categories c ON c.id=t.category_id "
        "LEFT JOIN providers pr ON pr.id=t.provider_id "
        "LEFT JOIN plan_items pi ON pi.id=t.plan_item_id "
        f"{where_sql} "
        "ORDER BY COALESCE(t.due_date, t.date) DESC, t.id DESC LIMIT 400",
        tuple(params),
    ).fetchall()

    total_income = total_expense = total_pending = total_received = total_paid_expense = total_overdue = 0
    income_by_pm_cents = {"cash": 0, "pix": 0, "card_credit": 0, "card_debit": 0, "transfer": 0, "other": 0}
    today = today_yyyy_mm_dd()
    ids = [r["id"] for r in rows]

    for r in rows:
        amount = int(r["final_amount_cents"] or r["amount_cents"] or 0)
        paid = int(r["paid_amount_cents"] or 0)
        balance = int(r["balance_cents"] or 0)
        if r["kind"] == "income":
            total_income += amount
            total_received += paid
        else:
            total_expense += amount
            total_paid_expense += paid
        if r["status"] == "pending":
            total_pending += balance
            if (r["due_date"] or r["date"]) < today:
                total_overdue += balance

    if ids:
        placeholders = ",".join("?" for _ in ids)
        pay_rows = db.execute(
            "SELECT p.payment_method, SUM(p.amount_cents) AS s "
            "FROM transaction_payments p JOIN transactions t ON t.id=p.transaction_id "
            f"WHERE t.kind='income' AND t.id IN ({placeholders}) GROUP BY p.payment_method",
            tuple(ids),
        ).fetchall()
        for r in pay_rows:
            pm = r["payment_method"] or "other"
            if pm == "card":
                pm = "card_credit"
            if pm not in income_by_pm_cents:
                pm = "other"
            income_by_pm_cents[pm] += int(r["s"] or 0)

    patients = db.execute("SELECT id, name FROM patients ORDER BY name ASC").fetchall()
    categories = db.execute("SELECT id, name FROM categories WHERE active=1 ORDER BY name ASC").fetchall()
    providers = db.execute("SELECT id, name FROM providers WHERE active=1 ORDER BY name ASC").fetchall()
    pm_labels = {k: v for k, v in PAYMENT_METHODS}

    income_by_pm = {k: cents_to_brl(v) for k, v in income_by_pm_cents.items()}
    totals = {
        "income": cents_to_brl(total_income),
        "expense": cents_to_brl(total_expense),
        "received": cents_to_brl(total_received),
        "paid_expense": cents_to_brl(total_paid_expense),
        "pending": cents_to_brl(total_pending),
        "overdue": cents_to_brl(total_overdue),
        "result": cents_to_brl(total_received - total_paid_expense),
    }
    return render_template(
        "transactions_list.html",
        rows=rows,
        cents_to_brl=cents_to_brl,
        pm_labels=pm_labels,
        filters=filters,
        totals=totals,
        income_by_pm=income_by_pm,
        patients=patients,
        categories=categories,
        providers=providers,
        pm=PAYMENT_METHODS,
        today=today,
    )


@bp.route("/transactions/<int:tid>/asaas", methods=["POST"])
@login_required
@finance_required
def transaction_asaas(tid):
    """Gera cobrança no Asaas quando ASAAS_API_KEY estiver configurada.

    Sem chave, mantém o botão ativo apenas como ponto preparado para vinculação.
    """
    db = get_db()
    row = db.execute(
        "SELECT t.*, p.name AS patient_name, p.phone AS patient_phone "
        "FROM transactions t LEFT JOIN patients p ON p.id=t.patient_id WHERE t.id=?",
        (tid,),
    ).fetchone()
    if not row:
        flash("Lançamento não encontrado.", "warning")
        return redirect(url_for("finance.transactions"))

    if row["kind"] != "income":
        flash("Boleto/ cobrança Asaas é usado apenas para entradas.", "warning")
        return redirect(url_for("finance.transactions"))

    api_key = (current_app.config.get("ASAAS_API_KEY") or os.environ.get("ASAAS_API_KEY") or app_setting("asaas_api_key", "") or "").strip()
    env = (current_app.config.get("ASAAS_ENV") or os.environ.get("ASAAS_ENV") or app_setting("asaas_env", "sandbox") or "sandbox").strip().lower()
    if not api_key:
        flash("Lugar do Asaas já está pronto. Falta colocar ASAAS_API_KEY nas variáveis do Coolify para emitir boleto/cobrança.", "info")
        return redirect(url_for("finance.transactions", status="pending"))

    base_url = "https://api.asaas.com/v3" if env == "production" else "https://sandbox.asaas.com/api/v3"
    name = row["patient_name"] or "Paciente sem cadastro"
    phone = "".join(ch for ch in (row["patient_phone"] or "") if ch.isdigit())
    amount_cents = int(row["balance_cents"] or row["final_amount_cents"] or row["amount_cents"] or 0)
    if amount_cents <= 0:
        flash("Esse lançamento não possui saldo para gerar cobrança.", "warning")
        return redirect(url_for("finance.transactions"))

    def post_json(endpoint, payload):
        data = json.dumps(payload).encode("utf-8")
        req = urlrequest.Request(
            base_url + endpoint,
            data=data,
            headers={"Content-Type": "application/json", "access_token": api_key, "User-Agent": "Eduarda-Imbelloni-Clinica/1.0"},
            method="POST",
        )
        with urlrequest.urlopen(req, timeout=25) as resp:
            return json.loads(resp.read().decode("utf-8"))

    try:
        customer_payload = {"name": name}
        if phone:
            customer_payload["mobilePhone"] = phone
        customer = post_json("/customers", customer_payload)
        billing_type = (request.form.get("billing_type") or "BOLETO").strip().upper()
        if billing_type not in {"BOLETO", "PIX", "CREDIT_CARD", "UNDEFINED"}:
            billing_type = "BOLETO"
        payment_payload = {
            "customer": customer.get("id"),
            "billingType": billing_type,
            "value": round(amount_cents / 100, 2),
            "dueDate": row["due_date"] or row["date"],
            "description": row["description"] or "Tratamento/atendimento clínico",
            "externalReference": f"transaction-{tid}",
        }
        payment = post_json("/payments", payment_payload)
        link = payment.get("invoiceUrl") or payment.get("bankSlipUrl") or payment.get("transactionReceiptUrl") or ""
        db.execute(
            "UPDATE transactions SET asaas_payment_id=?, asaas_invoice_url=?, asaas_status=?, asaas_payload=? WHERE id=?",
            (payment.get("id"), link, payment.get("status") or "created", json.dumps(payment, ensure_ascii=False)[:5000], tid),
        )
        db.commit()
        if link:
            flash(f"Cobrança Asaas criada e salva no lançamento. Link: {link}", "success")
        else:
            flash("Cobrança Asaas criada com sucesso. Confira no painel do Asaas.", "success")
    except HTTPError as e:
        try:
            detail = e.read().decode("utf-8")[:500]
        except Exception:
            detail = str(e)
        flash(f"Asaas retornou erro ao gerar cobrança: {detail}", "danger")
    except (URLError, TimeoutError, Exception) as e:
        flash(f"Não foi possível conectar ao Asaas agora: {e}", "danger")
    return redirect(url_for("finance.transactions", status="pending"))


@bp.route("/transactions/new", methods=["GET", "POST"])
@login_required
@finance_required
def transaction_new():
    db = get_db()
    data = get_common_form_data(db)

    if request.method == "POST":
        kind = request.form.get("kind", "income")
        status = request.form.get("status", "paid")
        date_eff = request.form.get("date", today_yyyy_mm_dd()) or today_yyyy_mm_dd()
        due_date = request.form.get("due_date", "").strip() or None
        gross_amount = parse_brl_to_cents(request.form.get("amount", "0"))
        discount_percent = parse_percent(request.form.get("discount_percent", "0"))
        discount_fixed = parse_brl_to_cents(request.form.get("discount_cents", "0"))
        payment_method = normalize_payment_method(request.form.get("payment_method", "pix"))
        description = request.form.get("description", "").strip()
        patient_id = request.form.get("patient_id", "").strip()
        category_id = request.form.get("category_id", "").strip()
        provider_id = request.form.get("provider_id", "").strip()
        plan_item_id = request.form.get("plan_item_id", "").strip()
        repasse_percent = request.form.get("repasse_percent", "").strip()
        installments_total = request.form.get("installments_total", "1").strip()

        if kind not in ("income", "expense"):
            kind = "income"
        if status not in ("paid", "pending"):
            status = "paid"
        pid = int(patient_id) if patient_id.isdigit() else None
        cid = int(category_id) if category_id.isdigit() else None
        prid = int(provider_id) if provider_id.isdigit() else None
        plan_iid = int(plan_item_id) if plan_item_id.isdigit() else None
        if plan_iid == 0:
            # 0 significa pagamento do plano/tratamento completo do paciente.
            if not pid:
                plan_iid = None
        elif plan_iid and pid:
            linked = db.execute("SELECT id FROM plan_items WHERE id=? AND patient_id=?", (plan_iid, pid)).fetchone()
            if not linked:
                plan_iid = None
        elif plan_iid:
            item = db.execute("SELECT patient_id FROM plan_items WHERE id=?", (plan_iid,)).fetchone()
            if item:
                pid = int(item["patient_id"])
            else:
                plan_iid = None

        if repasse_percent == "" and prid is not None:
            r = db.execute("SELECT default_repasse_percent FROM providers WHERE id=?", (prid,)).fetchone()
            repasse_percent = str(int(r["default_repasse_percent"] or 0)) if r else "0"
        try:
            repasse_percent_int = max(0, min(100, int(float((repasse_percent or "0").replace(",", ".")))))
        except Exception:
            repasse_percent_int = 0
        try:
            installments_total_int = max(1, min(36, int(installments_total or 1)))
        except Exception:
            installments_total_int = 1

        final_amount, total_discount = calc_final_amount(gross_amount, discount_percent, discount_fixed)
        if gross_amount <= 0 or final_amount <= 0:
            flash("Valor final precisa ser maior que zero.", "danger")
            tx_prefill = dict(request.form)
            tx_prefill["amount_brl"] = request.form.get("amount", "")
            tx_prefill["discount_cents_brl"] = request.form.get("discount_cents", "")
            return render_template("transaction_form.html", tx=tx_prefill, is_edit=False, **data)

        parent_id = None
        if installments_total_int > 1:
            base_due = due_date or date_eff
            base_desc = description or ("Entrada" if kind == "income" else "Saída")
            each = final_amount // installments_total_int
            remainder = final_amount % installments_total_int
            created = []
            for i in range(1, installments_total_int + 1):
                part_amount = each + (1 if i <= remainder else 0)
                part_due = add_months(base_due, i - 1)
                part_status = "pending"
                if status == "paid" and i == 1:
                    part_status = "paid"
                part_desc = f"{base_desc} - Parcela {i}/{installments_total_int}"
                tid = insert_transaction(
                    db,
                    kind=kind,
                    status=part_status,
                    date_eff=date_eff if part_status == "paid" else part_due,
                    due_date=part_due,
                    gross_amount=part_amount,
                    discount_percent=0,
                    discount_fixed=0,
                    description=part_desc,
                    pid=pid,
                    cid=cid,
                    prid=prid,
                    repasse_percent_int=repasse_percent_int,
                    payment_method=payment_method,
                    installments_total=installments_total_int,
                    installment_number=i,
                    parent_transaction_id=parent_id,
                    plan_item_id=plan_iid,
                )
                if parent_id is None:
                    parent_id = tid
                    db.execute("UPDATE transactions SET parent_transaction_id=? WHERE id=?", (parent_id, tid))
                else:
                    db.execute("UPDATE transactions SET parent_transaction_id=? WHERE id=?", (parent_id, tid))
                created.append(tid)
            db.commit()
            flash(f"Parcelamento criado com {installments_total_int} parcelas ✅", "success")
            return redirect(url_for("finance.transactions", status="pending"))

        insert_transaction(
            db,
            kind=kind,
            status=status,
            date_eff=date_eff,
            due_date=due_date,
            gross_amount=gross_amount,
            discount_percent=discount_percent,
            discount_fixed=discount_fixed,
            description=description,
            pid=pid,
            cid=cid,
            prid=prid,
            repasse_percent_int=repasse_percent_int,
            payment_method=payment_method,
            plan_item_id=plan_iid,
        )
        db.commit()
        flash("Lançamento salvo ✅", "success")
        return redirect(url_for("finance.transactions"))

    tx_prefill = {
        "kind": request.args.get("kind", "income"),
        "status": request.args.get("status", "paid"),
        "date": request.args.get("date", today_yyyy_mm_dd()),
        "due_date": request.args.get("due_date", ""),
        "payment_method": request.args.get("payment_method", "pix"),
        "installments_total": request.args.get("installments_total", 1),
        "discount_percent": 0,
        "discount_cents_brl": "",
        "patient_id": request.args.get("patient_id", ""),
        "plan_item_id": request.args.get("plan_item_id", ""),
        "description": request.args.get("description", ""),
    }

    # Se vier do Plano/Ficha, já puxa descrição, valor e paciente.
    plan_item_arg = request.args.get("plan_item_id", "").strip()
    patient_arg = request.args.get("patient_id", "").strip()
    if plan_item_arg in {"0", "geral", "plano"} and patient_arg.isdigit():
        total_row = db.execute("SELECT COALESCE(SUM(amount_cents),0) AS total FROM plan_items WHERE patient_id=?", (int(patient_arg),)).fetchone()
        tx_prefill["patient_id"] = int(patient_arg)
        tx_prefill["plan_item_id"] = 0
        tx_prefill["description"] = request.args.get("description", "Pagamento do tratamento completo")
        tx_prefill["amount_brl"] = cents_to_brl(int(total_row["total"] or 0))
        tx_prefill["status"] = request.args.get("status", "paid")
    elif plan_item_arg.isdigit():
        item = db.execute("SELECT * FROM plan_items WHERE id=?", (int(plan_item_arg),)).fetchone()
        if item:
            tx_prefill["patient_id"] = item["patient_id"]
            tx_prefill["plan_item_id"] = item["id"]
            tx_prefill["description"] = item["procedure"]
            tx_prefill["amount_brl"] = cents_to_brl(int(item["amount_cents"] or 0))
            tx_prefill["status"] = "pending"

    return render_template("transaction_form.html", tx=tx_prefill, is_edit=False, **data)


@bp.route("/transactions/<int:tid>/edit", methods=["GET", "POST"])
@login_required
@finance_required
def transaction_edit(tid: int):
    db = get_db()
    tx = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    if not tx:
        flash("Lançamento não encontrado.", "danger")
        return redirect(url_for("finance.transactions"))
    data = get_common_form_data(db)

    if request.method == "POST":
        kind = request.form.get("kind", tx["kind"])
        status = request.form.get("status", tx["status"])
        date_eff = request.form.get("date", tx["date"]) or today_yyyy_mm_dd()
        due_date = request.form.get("due_date", "").strip() or None
        gross_amount = parse_brl_to_cents(request.form.get("amount", "0"))
        discount_percent = parse_percent(request.form.get("discount_percent", "0"))
        discount_fixed = parse_brl_to_cents(request.form.get("discount_cents", "0"))
        payment_method = normalize_payment_method(request.form.get("payment_method", tx["payment_method"]))
        description = request.form.get("description", "").strip()
        patient_id = request.form.get("patient_id", "").strip()
        category_id = request.form.get("category_id", "").strip()
        provider_id = request.form.get("provider_id", "").strip()
        plan_item_id = request.form.get("plan_item_id", "").strip()
        repasse_percent = request.form.get("repasse_percent", "").strip()

        pid = int(patient_id) if patient_id.isdigit() else None
        cid = int(category_id) if category_id.isdigit() else None
        prid = int(provider_id) if provider_id.isdigit() else None
        plan_iid = int(plan_item_id) if plan_item_id.isdigit() else None

        # Garante que o tratamento vinculado pertence ao paciente escolhido.
        # 0 significa plano/tratamento completo do paciente, não um procedimento específico.
        if plan_iid == 0:
            if not pid:
                plan_iid = None
        elif plan_iid and pid:
            linked = db.execute("SELECT id FROM plan_items WHERE id=? AND patient_id=?", (plan_iid, pid)).fetchone()
            if not linked:
                plan_iid = None
        elif plan_iid:
            item = db.execute("SELECT patient_id FROM plan_items WHERE id=?", (plan_iid,)).fetchone()
            if item:
                pid = int(item["patient_id"])
            else:
                plan_iid = None

        try:
            repasse_percent_int = max(0, min(100, int(float((repasse_percent or "0").replace(",", ".")))))
        except Exception:
            repasse_percent_int = 0

        final_amount, total_discount = calc_final_amount(gross_amount, discount_percent, discount_fixed)
        if gross_amount <= 0 or final_amount <= 0:
            flash("Valor final precisa ser maior que zero.", "danger")
            tx_dict = dict(tx)
            tx_dict.update(request.form)
            tx_dict["amount_brl"] = request.form.get("amount", "")
            tx_dict["discount_cents_brl"] = request.form.get("discount_cents", "")
            return render_template("transaction_form.html", tx=tx_dict, is_edit=True, **data)

        db.execute(
            """
            UPDATE transactions
            SET kind=?, date=?, due_date=?, amount_cents=?, payment_method=?, description=?, patient_id=?, category_id=?, provider_id=?,
                repasse_percent=?, gross_amount_cents=?, discount_percent=?, discount_cents=?, final_amount_cents=?, plan_item_id=?
            WHERE id=?
            """,
            (kind, date_eff, due_date, final_amount, payment_method, description, pid, cid, prid, repasse_percent_int, gross_amount, discount_percent, total_discount, final_amount, plan_iid, tid),
        )
        sync_transaction_payments(db, tid)
        tx2 = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
        if status == "paid" and int(tx2["balance_cents"] or 0) > 0:
            add_payment(db, tid, int(tx2["balance_cents"] or 0), payment_method, date_eff, "Complemento ao editar")
        db.commit()
        flash("Lançamento atualizado ✅", "success")
        return redirect(url_for("finance.transactions"))

    tx_dict = dict(tx)
    if "plan_item_id" not in tx_dict:
        tx_dict["plan_item_id"] = None
    tx_dict["amount_brl"] = cents_to_brl(int(tx["gross_amount_cents"] if tx["gross_amount_cents"] is not None else tx["amount_cents"] or 0))
    tx_dict["discount_cents_brl"] = cents_to_brl(int(tx["discount_cents"] or 0))
    tx_dict["discount_percent_str"] = fmt_percent(tx["discount_percent"])
    return render_template("transaction_form.html", tx=tx_dict, is_edit=True, **data)


@bp.route("/transactions/<int:tid>/delete", methods=["POST"])
@login_required
@finance_required
def transaction_delete(tid: int):
    db = get_db()
    db.execute("DELETE FROM transactions WHERE id=?", (tid,))
    db.commit()
    flash("Lançamento removido.", "info")
    return redirect(url_for("finance.transactions"))


@bp.route("/transactions/<int:tid>/settle", methods=["POST"])
@login_required
@finance_required
def transaction_settle(tid: int):
    db = get_db()
    tx = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    if not tx:
        flash("Lançamento não encontrado.", "danger")
        return redirect(url_for("finance.transactions"))
    remaining = int(tx["balance_cents"] or 0)
    if remaining <= 0:
        flash("Lançamento já está pago.", "info")
        return redirect(url_for("finance.transactions"))
    payment_method = normalize_payment_method(request.form.get("payment_method", "pix"))
    date_eff = request.form.get("date", today_yyyy_mm_dd())
    add_payment(db, tid, remaining, payment_method, date_eff, "Baixa rápida")
    db.commit()
    flash("Baixado ✅", "success")
    return redirect(url_for("finance.transactions"))


@bp.route("/transactions/<int:tid>/payments", methods=["GET", "POST"])
@login_required
@finance_required
def transaction_payments(tid: int):
    db = get_db()
    tx = db.execute(
        "SELECT t.*, p.name AS patient_name, c.name AS category_name, pr.name AS provider_name "
        "FROM transactions t "
        "LEFT JOIN patients p ON p.id=t.patient_id "
        "LEFT JOIN categories c ON c.id=t.category_id "
        "LEFT JOIN providers pr ON pr.id=t.provider_id "
        "WHERE t.id=?",
        (tid,),
    ).fetchone()
    if not tx:
        flash("Lançamento não encontrado.", "danger")
        return redirect(url_for("finance.transactions"))

    if request.method == "POST":
        amount = parse_brl_to_cents(request.form.get("amount", "0"))
        payment_method = normalize_payment_method(request.form.get("payment_method", "pix"))
        pay_date = request.form.get("date", today_yyyy_mm_dd()) or today_yyyy_mm_dd()
        notes = request.form.get("notes", "")
        if add_payment(db, tid, amount, payment_method, pay_date, notes):
            db.commit()
            flash("Pagamento registrado ✅", "success")
        else:
            flash("Pagamento inválido.", "danger")
        return redirect(url_for("finance.transaction_payments", tid=tid))

    sync_transaction_payments(db, tid)
    db.commit()
    tx = db.execute(
        "SELECT t.*, p.name AS patient_name, c.name AS category_name, pr.name AS provider_name "
        "FROM transactions t "
        "LEFT JOIN patients p ON p.id=t.patient_id "
        "LEFT JOIN categories c ON c.id=t.category_id "
        "LEFT JOIN providers pr ON pr.id=t.provider_id "
        "WHERE t.id=?",
        (tid,),
    ).fetchone()
    payments = db.execute("SELECT * FROM transaction_payments WHERE transaction_id=? ORDER BY date DESC, id DESC", (tid,)).fetchall()
    pm_labels = {k: v for k, v in PAYMENT_METHODS}
    return render_template("transaction_payments.html", tx=tx, payments=payments, pm=PAYMENT_METHODS, pm_labels=pm_labels, cents_to_brl=cents_to_brl, today=today_yyyy_mm_dd())


@bp.route("/payments/<int:pid>/delete", methods=["POST"])
@login_required
@finance_required
def payment_delete(pid: int):
    db = get_db()
    payment = db.execute("SELECT * FROM transaction_payments WHERE id=?", (pid,)).fetchone()
    if not payment:
        flash("Pagamento não encontrado.", "danger")
        return redirect(url_for("finance.transactions"))
    tid = int(payment["transaction_id"])
    db.execute("DELETE FROM transaction_payments WHERE id=?", (pid,))
    sync_transaction_payments(db, tid)
    db.commit()
    flash("Pagamento removido.", "info")
    return redirect(url_for("finance.transaction_payments", tid=tid))


@bp.route("/reports")
@login_required
@finance_required
def reports():
    db = get_db()
    today = date.today()
    default_from = today.replace(day=1).isoformat()
    default_to = today.isoformat()
    date_from = request.args.get("from", default_from).strip() or default_from
    date_to = request.args.get("to", default_to).strip() or default_to

    summary = db.execute(
        """
        SELECT
          SUM(CASE WHEN kind='income' THEN final_amount_cents ELSE 0 END) AS income_total,
          SUM(CASE WHEN kind='income' THEN paid_amount_cents ELSE 0 END) AS income_paid,
          SUM(CASE WHEN kind='income' THEN balance_cents ELSE 0 END) AS income_pending,
          SUM(CASE WHEN kind='expense' THEN final_amount_cents ELSE 0 END) AS expense_total,
          SUM(CASE WHEN kind='expense' THEN paid_amount_cents ELSE 0 END) AS expense_paid,
          SUM(CASE WHEN status='pending' AND COALESCE(due_date,date) < date('now') THEN balance_cents ELSE 0 END) AS overdue
        FROM transactions
        WHERE COALESCE(due_date, date) BETWEEN ? AND ?
        """,
        (date_from, date_to),
    ).fetchone()

    by_method = db.execute(
        """
        SELECT p.payment_method, SUM(CASE WHEN t.kind='income' THEN p.amount_cents ELSE 0 END) AS income,
               SUM(CASE WHEN t.kind='expense' THEN p.amount_cents ELSE 0 END) AS expense
        FROM transaction_payments p
        JOIN transactions t ON t.id=p.transaction_id
        WHERE p.date BETWEEN ? AND ?
        GROUP BY p.payment_method
        ORDER BY p.payment_method
        """,
        (date_from, date_to),
    ).fetchall()

    by_category = db.execute(
        """
        SELECT COALESCE(c.name, 'Sem categoria') AS name,
               SUM(CASE WHEN t.kind='income' THEN t.final_amount_cents ELSE 0 END) AS income,
               SUM(CASE WHEN t.kind='expense' THEN t.final_amount_cents ELSE 0 END) AS expense
        FROM transactions t
        LEFT JOIN categories c ON c.id=t.category_id
        WHERE COALESCE(t.due_date, t.date) BETWEEN ? AND ?
        GROUP BY COALESCE(c.name, 'Sem categoria')
        ORDER BY income DESC, expense DESC
        """,
        (date_from, date_to),
    ).fetchall()

    by_provider = db.execute(
        """
        SELECT COALESCE(pr.name, 'Sem profissional') AS name,
               SUM(CASE WHEN t.kind='income' THEN t.final_amount_cents ELSE 0 END) AS produced,
               SUM(CASE WHEN t.kind='income' THEN (t.final_amount_cents * t.repasse_percent)/100 ELSE 0 END) AS repasse
        FROM transactions t
        LEFT JOIN providers pr ON pr.id=t.provider_id
        WHERE COALESCE(t.due_date, t.date) BETWEEN ? AND ?
        GROUP BY COALESCE(pr.name, 'Sem profissional')
        ORDER BY produced DESC
        """,
        (date_from, date_to),
    ).fetchall()

    upcoming = db.execute(
        """
        SELECT t.*, p.name AS patient_name
        FROM transactions t
        LEFT JOIN patients p ON p.id=t.patient_id
        WHERE t.status='pending'
          AND COALESCE(t.due_date, t.date) BETWEEN date('now') AND date('now', '+90 day')
        ORDER BY COALESCE(t.due_date, t.date) ASC, t.id ASC
        LIMIT 80
        """
    ).fetchall()

    return render_template(
        "finance_reports.html",
        date_from=date_from,
        date_to=date_to,
        summary=summary,
        by_method=by_method,
        by_category=by_category,
        by_provider=by_provider,
        upcoming=upcoming,
        cents_to_brl=cents_to_brl,
        pm_labels={k: v for k, v in PAYMENT_METHODS},
    )


def fetch_export_rows(db, filters):
    where_sql, params = build_transactions_query(filters)
    return db.execute(
        "SELECT t.*, p.name AS patient_name, c.name AS category_name, pr.name AS provider_name "
        "FROM transactions t "
        "LEFT JOIN patients p ON p.id=t.patient_id "
        "LEFT JOIN categories c ON c.id=t.category_id "
        "LEFT JOIN providers pr ON pr.id=t.provider_id "
        f"{where_sql} "
        "ORDER BY COALESCE(t.due_date, t.date) DESC, t.id DESC",
        tuple(params),
    ).fetchall()


@bp.route("/export/csv")
@login_required
@finance_required
def export_csv():
    db = get_db()
    rows = fetch_export_rows(db, get_filters())
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["ID", "Tipo", "Status", "Data", "Vencimento", "Paciente", "Categoria", "Profissional", "Descrição", "Valor bruto", "Desconto", "Valor final", "Pago", "Saldo", "Pagamento", "Parcelas", "Repasse %"])
    for r in rows:
        writer.writerow([
            r["id"], "Entrada" if r["kind"] == "income" else "Saída", "Pago" if r["status"] == "paid" else "Pendente", r["date"], r["due_date"] or "",
            r["patient_name"] or "", r["category_name"] or "", r["provider_name"] or "", r["description"] or "",
            cents_to_brl(r["gross_amount_cents"] or r["amount_cents"]), cents_to_brl(r["discount_cents"] or 0), cents_to_brl(r["final_amount_cents"] or r["amount_cents"]),
            cents_to_brl(r["paid_amount_cents"] or 0), cents_to_brl(r["balance_cents"] or 0), normalize_payment_method(r["payment_method"]), f"{r['installment_number']}/{r['installments_total']}", r["repasse_percent"] or 0,
        ])
    return Response(output.getvalue().encode("utf-8-sig"), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=financeiro_transacoes.csv"})


@bp.route("/export/excel")
@login_required
@finance_required
def export_excel():
    db = get_db()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        flash("Biblioteca openpyxl não instalada. Rode: pip install openpyxl", "danger")
        return redirect(url_for("finance.transactions"))

    rows = fetch_export_rows(db, get_filters())
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações"
    headers = ["ID", "Tipo", "Status", "Data", "Vencimento", "Paciente", "Categoria", "Profissional", "Descrição", "Valor bruto", "Desconto", "Valor final", "Pago", "Saldo", "Forma", "Parcela", "Repasse %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF3FF")
    for r in rows:
        ws.append([
            r["id"], "Entrada" if r["kind"] == "income" else "Saída", "Pago" if r["status"] == "paid" else "Pendente", r["date"], r["due_date"] or "",
            r["patient_name"] or "", r["category_name"] or "", r["provider_name"] or "", r["description"] or "",
            (r["gross_amount_cents"] or r["amount_cents"] or 0) / 100, (r["discount_cents"] or 0) / 100, (r["final_amount_cents"] or r["amount_cents"] or 0) / 100,
            (r["paid_amount_cents"] or 0) / 100, (r["balance_cents"] or 0) / 100, r["payment_method"], f"{r['installment_number']}/{r['installments_total']}", r["repasse_percent"] or 0,
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 40)

    wp = wb.create_sheet("Pagamentos")
    wp.append(["ID", "Transação", "Data", "Valor", "Forma", "Observação"])
    for cell in wp[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAF3FF")
    pays = db.execute("SELECT * FROM transaction_payments ORDER BY date DESC, id DESC").fetchall()
    for p in pays:
        wp.append([p["id"], p["transaction_id"], p["date"], (p["amount_cents"] or 0) / 100, p["payment_method"], p["notes"] or ""])
    for col in wp.columns:
        wp.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col) + 2, 40)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="financeiro_eduarda_imbelloni.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/export/pdf")
@login_required
@finance_required
def export_pdf():
    db = get_db()
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except Exception:
        flash("Biblioteca reportlab não instalada. Rode: pip install reportlab", "danger")
        return redirect(url_for("finance.transactions"))

    filters = get_filters()
    rows = fetch_export_rows(db, filters)
    total_income = sum(int(r["paid_amount_cents"] or 0) for r in rows if r["kind"] == "income")
    total_expense = sum(int(r["paid_amount_cents"] or 0) for r in rows if r["kind"] == "expense")
    pending = sum(int(r["balance_cents"] or 0) for r in rows if r["status"] == "pending")

    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = [Paragraph("Relatório Financeiro - Eduarda Imbelloni", styles["Title"]), Spacer(1, 12)]
    story.append(Paragraph(f"Recebido: R$ {cents_to_brl(total_income)} | Saídas pagas: R$ {cents_to_brl(total_expense)} | Resultado: R$ {cents_to_brl(total_income-total_expense)} | Pendente: R$ {cents_to_brl(pending)}", styles["Normal"]))
    story.append(Spacer(1, 12))
    data = [["Data", "Tipo", "Status", "Paciente", "Descrição", "Final", "Pago", "Saldo"]]
    for r in rows[:80]:
        data.append([r["date"], "Ent." if r["kind"] == "income" else "Saí.", "Pago" if r["status"] == "paid" else "Pend.", (r["patient_name"] or "")[:16], (r["description"] or "")[:24], cents_to_brl(r["final_amount_cents"] or r["amount_cents"]), cents_to_brl(r["paid_amount_cents"] or 0), cents_to_brl(r["balance_cents"] or 0)])
    table = Table(data, repeatRows=1, colWidths=[58, 36, 44, 72, 120, 55, 55, 55])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF3FF")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    doc.build(story)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="relatorio_financeiro.pdf", mimetype="application/pdf")


# Categorias
@bp.route("/categories")
@login_required
@finance_required
def categories_list():
    db = get_db()
    rows = db.execute("SELECT * FROM categories ORDER BY active DESC, name ASC").fetchall()
    return render_template("categories_list.html", rows=rows)


@bp.route("/categories/new", methods=["GET", "POST"])
@login_required
@finance_required
def category_new():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        kind = request.form.get("kind", "both")
        if not name:
            flash("Nome é obrigatório.", "danger")
            return render_template("category_form.html", c=None)
        db.execute("INSERT OR IGNORE INTO categories(name, kind) VALUES(?,?)", (name, kind))
        db.commit()
        flash("Categoria salva ✅", "success")
        return redirect(url_for("finance.categories_list"))
    return render_template("category_form.html", c=None)


@bp.route("/categories/<int:cid>/toggle", methods=["POST"])
@login_required
@finance_required
def category_toggle(cid: int):
    db = get_db()
    row = db.execute("SELECT active FROM categories WHERE id=?", (cid,)).fetchone()
    if row:
        newv = 0 if int(row["active"]) == 1 else 1
        db.execute("UPDATE categories SET active=? WHERE id=?", (newv, cid))
        db.commit()
    return redirect(url_for("finance.categories_list"))


# Caixa
@bp.route("/caixa", methods=["GET", "POST"])
@login_required
@finance_required
def caixa():
    db = get_db()
    open_cash_id = get_open_cash_session_id()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "open" and not open_cash_id:
            open_balance = parse_brl_to_cents(request.form.get("open_balance", "0"))
            notes = request.form.get("notes", "").strip()
            db.execute("INSERT INTO cash_sessions(opened_at, open_balance_cents, notes) VALUES(datetime('now'), ?, ?)", (open_balance, notes))
            db.commit()
            flash("Caixa aberto ✅", "success")
            return redirect(url_for("finance.caixa"))
        if action == "close" and open_cash_id:
            close_balance = parse_brl_to_cents(request.form.get("close_balance", "0"))
            notes = request.form.get("notes", "").strip()
            row_open = db.execute("SELECT open_balance_cents FROM cash_sessions WHERE id=?", (open_cash_id,)).fetchone()
            open_balance = int(row_open["open_balance_cents"] or 0)
            rows = db.execute(
                "SELECT t.kind, SUM(p.amount_cents) s FROM transaction_payments p JOIN transactions t ON t.id=p.transaction_id "
                "WHERE p.payment_method='cash' AND p.cash_session_id=? GROUP BY t.kind",
                (open_cash_id,),
            ).fetchall()
            incomes = expenses = 0
            for r in rows:
                if r["kind"] == "income":
                    incomes = int(r["s"] or 0)
                else:
                    expenses = int(r["s"] or 0)
            expected = open_balance + incomes - expenses
            db.execute(
                "UPDATE cash_sessions SET closed_at=datetime('now'), close_balance_cents=?, expected_balance_cents=?, notes=COALESCE(notes,'') || CASE WHEN ?<>'' THEN char(10)||? ELSE '' END WHERE id=?",
                (close_balance, expected, notes, notes, open_cash_id),
            )
            db.commit()
            flash(f"Caixa fechado. Esperado: {cents_to_brl(expected)} | Informado: {cents_to_brl(close_balance)}", "info")
            return redirect(url_for("finance.caixa_history"))

    session_row = None
    expected_now = None
    if open_cash_id:
        session_row = db.execute("SELECT * FROM cash_sessions WHERE id=?", (open_cash_id,)).fetchone()
        open_balance = int(session_row["open_balance_cents"] or 0)
        rows = db.execute(
            "SELECT t.kind, SUM(p.amount_cents) s FROM transaction_payments p JOIN transactions t ON t.id=p.transaction_id "
            "WHERE p.payment_method='cash' AND p.cash_session_id=? GROUP BY t.kind",
            (open_cash_id,),
        ).fetchall()
        incomes = expenses = 0
        for r in rows:
            if r["kind"] == "income":
                incomes = int(r["s"] or 0)
            else:
                expenses = int(r["s"] or 0)
        expected_now = open_balance + incomes - expenses
    return render_template("caixa.html", open_cash_id=open_cash_id, session_row=session_row, expected_now=expected_now, cents_to_brl=cents_to_brl)


@bp.route("/caixa/historico")
@login_required
@finance_required
def caixa_history():
    db = get_db()
    rows = db.execute("SELECT * FROM cash_sessions ORDER BY id DESC LIMIT 60").fetchall()
    return render_template("caixa_historico.html", rows=rows, cents_to_brl=cents_to_brl)


# Profissionais (Dentistas) + Repasses
@bp.route("/providers")
@login_required
@finance_required
def providers_list():
    db = get_db()
    rows = db.execute("SELECT * FROM providers ORDER BY active DESC, name ASC").fetchall()
    return render_template("providers_list.html", rows=rows)


@bp.route("/providers/new", methods=["GET", "POST"])
@login_required
@finance_required
def provider_new():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role = request.form.get("role", "Dentista").strip() or "Dentista"
        default_repasse_percent = request.form.get("default_repasse_percent", "0").strip()
        try:
            drp = max(0, min(100, int(float(default_repasse_percent.replace(",", ".")))))
        except Exception:
            drp = 0
        if not name:
            flash("Nome é obrigatório.", "danger")
            return render_template("provider_form.html", p=None)
        db.execute("INSERT INTO providers(name, role, default_repasse_percent) VALUES(?,?,?)", (name, role, drp))
        db.commit()
        flash("Profissional salvo ✅", "success")
        return redirect(url_for("finance.providers_list"))
    return render_template("provider_form.html", p=None)


@bp.route("/providers/<int:pid>/toggle", methods=["POST"])
@login_required
@finance_required
def provider_toggle(pid: int):
    db = get_db()
    row = db.execute("SELECT active FROM providers WHERE id=?", (pid,)).fetchone()
    if row:
        newv = 0 if int(row["active"]) == 1 else 1
        db.execute("UPDATE providers SET active=? WHERE id=?", (newv, pid))
        db.commit()
    return redirect(url_for("finance.providers_list"))


@bp.route("/repasses")
@login_required
@finance_required
def repasses():
    db = get_db()
    month_start = date.today().replace(day=1).isoformat()
    date_from = request.args.get("from", month_start).strip() or month_start
    date_to = request.args.get("to", date.today().isoformat()).strip() or date.today().isoformat()
    provider_id = request.args.get("provider_id", "").strip()
    repasse_status = request.args.get("repasse_status", "pending").strip()

    where = ["t.kind='income'", "t.status='paid'", "t.repasse_percent>0", "t.date BETWEEN ? AND ?"]
    params = [date_from, date_to]
    if provider_id.isdigit():
        where.append("t.provider_id=?")
        params.append(int(provider_id))
    if repasse_status == "paid":
        where.append("t.repasse_paid=1")
    elif repasse_status == "all":
        pass
    else:
        repasse_status = "pending"
        where.append("t.repasse_paid=0")
    where_sql = "WHERE " + " AND ".join(where)

    rows = db.execute(
        "SELECT pr.id, pr.name, "
        "SUM(CASE WHEN t.status='paid' AND t.date BETWEEN ? AND ? THEN (t.final_amount_cents*t.repasse_percent)/100 ELSE 0 END) AS repasse_periodo, "
        "SUM(CASE WHEN t.status='paid' AND t.repasse_paid=0 THEN (t.final_amount_cents*t.repasse_percent)/100 ELSE 0 END) AS repasse_pendente "
        "FROM providers pr "
        "LEFT JOIN transactions t ON t.provider_id=pr.id AND t.kind='income' "
        "WHERE pr.active=1 "
        "GROUP BY pr.id, pr.name "
        "ORDER BY pr.name ASC",
        (date_from, date_to),
    ).fetchall()

    detail = db.execute(
        "SELECT t.*, pr.name AS provider_name, p.name AS patient_name "
        "FROM transactions t "
        "JOIN providers pr ON pr.id=t.provider_id "
        "LEFT JOIN patients p ON p.id=t.patient_id "
        f"{where_sql} "
        "ORDER BY t.date DESC, t.id DESC LIMIT 300",
        tuple(params),
    ).fetchall()
    providers = db.execute("SELECT id, name FROM providers WHERE active=1 ORDER BY name ASC").fetchall()
    total_repasse = sum(int((int(t["final_amount_cents"] or t["amount_cents"] or 0) * int(t["repasse_percent"] or 0)) // 100) for t in detail)
    return render_template("repasses.html", rows=rows, pend=detail, providers=providers, filters={"from": date_from, "to": date_to, "provider_id": provider_id, "repasse_status": repasse_status}, total_repasse=total_repasse, cents_to_brl=cents_to_brl)


@bp.route("/repasses/<int:tid>/pay", methods=["POST"])
@login_required
@finance_required
def repasse_pay(tid: int):
    db = get_db()
    note = request.form.get("note", "").strip()
    db.execute("UPDATE transactions SET repasse_paid=1, repasse_paid_at=datetime('now'), repasse_payment_note=? WHERE id=?", (note or None, tid))
    db.commit()
    flash("Repasse marcado como pago ✅", "success")
    return redirect(request.referrer or url_for("finance.repasses"))


@bp.route("/repasses/<int:tid>/unpay", methods=["POST"])
@login_required
@finance_required
def repasse_unpay(tid: int):
    db = get_db()
    db.execute("UPDATE transactions SET repasse_paid=0, repasse_paid_at=NULL, repasse_payment_note=NULL WHERE id=?", (tid,))
    db.commit()
    flash("Repasse voltou para pendente.", "info")
    return redirect(request.referrer or url_for("finance.repasses"))
